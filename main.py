import json
from openpyxl import load_workbook
from rich.console import Console
import requests
from selenium import webdriver
import selenium
import selenium.common
import selenium.types
import selenium.webdriver
from selenium.webdriver.common.by import By

class TempWebElement:
    def __init__(self) -> None:
        self.text = ""

class ErrorChecker:
    def __init__(self) -> None:
        self.driver = webdriver.Firefox()
    
    def check_error(self, url: str) -> bool:
        self.driver.get(url)

        self.driver.implicitly_wait(2)
        try:
            error = self.driver.find_element(by=By.CLASS_NAME, value="filename")
        except:
            console.log(f"[red]Error occurred with {name}[/red]")
            error = TempWebElement()

        return error.text != ""

    def end(self):
        self.driver.quit()

console = Console()
error_checker = ErrorChecker()

workbook = load_workbook(filename="./movies.xlsx")
sheet = workbook.active

working = []

for value in sheet.iter_rows(min_row=200, max_row=300, min_col=1, max_col=1, values_only=True):
    value_split = value[0].split(" ")
    name = " ".join(value_split[:-1])
    link = value_split[-1]
    console.log(f"Testing [bright_magenta]{name}[/bright_magenta] at [bright_cyan bold]{link}[/bright_cyan bold]")
    
    if error_checker.check_error(f"{link}"):
        working.append(name)
        print(len(working))
        console.log(f"[bright_green]{name} works[/bright_green]")

    print("")

error_checker.driver.close()
print(", ".join(working))
open("working.json", "w")
json.dump(working, open("working.json", "w"))

# error_checker.check_error("https://mega.nz/file/76YAASwR#3EUaw1kd_mqrCpDBKg21EpFrBIowQwa3mzEbj9bZDl8")
# error_checker.check_error("https://mega.nz/file/SqoDAIZR#ctCRL6U1OlrlkA_KRagX1EgJqQAN1sw-0m9")

